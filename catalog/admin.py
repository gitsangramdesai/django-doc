from django.contrib import admin
from .models import MyModel
from django.http import HttpResponse

# ... export functions will go here ...

class MyModelAdmin(admin.ModelAdmin):
    def export_csv(modeladmin, request, queryset):
        import csv
        from django.utils.encoding import smart_str

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename=mymodel.csv'
        writer = csv.writer(response, csv.excel)
        response.write(u'\ufeff'.encode('utf8'))  # BOM (optional...Excel needs it to open UTF-8 file properly)
        writer.writerow([
            smart_str(u"ID"),
            smart_str(u"Title"),
            smart_str(u"Description"),
        ])
        for obj in queryset:
            writer.writerow([
                smart_str(obj.pk),
                smart_str(obj.title),
                smart_str(obj.description),
            ])
        return response
    export_csv.short_description = u"Export CSV"

    def export_xls(modeladmin, request, queryset):
        import xlwt
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename=mymodel.xls'

        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet("MyModel")

        row_num = 0

        columns = [
            (u"ID", 2000),
            (u"Title", 6000),
            (u"Description", 8000),
        ]

        font_style = xlwt.XFStyle()
        font_style.font.bold = True

        for col_num in range(len(columns)):
            ws.write(row_num, col_num, columns[col_num][0], font_style)
            # set column width
            ws.col(col_num).width = columns[col_num][1]

        font_style = xlwt.XFStyle()
        font_style.alignment.wrap = 1

        for obj in queryset:
            row_num += 1
            row = [
                obj.pk,
                obj.title,
                obj.description,
            ]
            for col_num in range(len(row)):
                ws.write(row_num, col_num, row[col_num], font_style)

        wb.save(response)
        return response
    export_xls.short_description = u"Export XLS"


    def export_xlsx(modeladmin, request, queryset):
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font,Color

        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename=mymodel.xls'

        wb = Workbook()

        # grab the active worksheet
        ws = wb.active

        # Rows can also be appended
        ws.append(["ID", "Title", "Description"])

        for obj in queryset:
            row = [
                obj.pk,
                obj.title,
                obj.description,
            ]
            ws.append(row)

        header_row = ws.row_dimensions[1]
        header_row.font =Font(bold=True)

        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        redFill = PatternFill(start_color='FFFF0000',
                              end_color='FFFF0000',
                              fill_type='solid')

        FreeMono_font = Font(name='FreeMono',
                    size=11,
                    bold=False,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FF000000')


        ws.cell(row=1, column=1).border = thin_border
        ws.cell(row=1, column=2).border = thin_border
        ws.cell(row=1, column=3).border = thin_border

        ws.cell(row=1, column=1).fill = redFill
        ws.cell(row=1, column=2).fill = redFill
        ws.cell(row=1, column=3).fill = redFill

        ws.cell(row=1, column=1).font = FreeMono_font
        ws.cell(row=1, column=2).font = FreeMono_font
        ws.cell(row=1, column=3).font = FreeMono_font


        wb.save(response)
        return response

    export_xlsx.short_description = u"Export XLSX"

    actions = [export_csv, export_xls, export_xlsx]

admin.site.register(MyModel, MyModelAdmin)
